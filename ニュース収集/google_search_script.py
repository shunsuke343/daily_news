import sys
sys.stdout.reconfigure(encoding='utf-8')

import os
import time
import pandas as pd
import requests
import feedparser
from bs4 import BeautifulSoup
from urllib.parse import urlparse, quote, urljoin, parse_qs
from concurrent.futures import ThreadPoolExecutor, as_completed
from datetime import datetime, timedelta
from dateutil import parser
import difflib
import json
import csv
from openpyxl import load_workbook
from collections import Counter
import math
import re
import base64
from pathlib import Path
from io import BytesIO

# --- Settings paths ---
API_KEYS_PATH = Path("api_keys.json")
DEPT_SETTINGS_PATH = Path("department_settings.json")
PROMPT_PATH = Path("\u30d7\u30ed\u30f3\u30d7\u30c8.md")
PROMPT_TEXT = ""

# Defaults (can be overridden by department_settings.json)
RSS_FEEDS = []
COUNTRY_SETTINGS = {}
INTERIOR_KEYWORDS = []
INTERIOR_KEYWORDS_LOWER = []
SYNONYM_GROUPS = []
SUBJECT_NAME = "\u5185\u88c5"
PHOTO_TARGET_LABEL = "\u8eca\u5185\u88c5"
DEPARTMENT = os.environ.get("DEPARTMENT", "")
TARGET_DATE_ONLY_YESTERDAY = True

# Output files
EXCEL_FILE = "search_results.csv"
LEGACY_EXCEL_FILE = "search_results.xlsx"

# Output columns
OUTPUT_COLUMNS = [
    "\u56fd",
    "\u691c\u7d22\u30ef\u30fc\u30c9",
    "\u30bf\u30a4\u30c8\u30eb",
    "\u30bf\u30a4\u30c8\u30eb\uff08\u65e5\u672c\u8a9e\uff09",
    "\u65e5\u4ed8",
    "\u5185\u5bb9",
    "\u5185\u5bb9\uff08\u65e5\u672c\u8a9e\uff09",
    "\u95a2\u9023\u5ea6",
    "\u95a2\u9023\u5ea6\u30b9\u30b3\u30a2",
    "\u95a2\u9023\u30ad\u30fc\u30ef\u30fc\u30c9",
    "\u51fa\u5c55\u30b5\u30a4\u30c8",
    "\u753b\u50cfURL",
    "URL",
    "\u30bd\u30fc\u30b9",
    "HTML\u53d6\u5f97",
    "LLM\u5224\u5b9a",
    "\u753b\u50cf\u5224\u5b9a",
    "LLM\u5f8c\u51e6\u7406",
    "\u30b9\u30c6\u30fc\u30bf\u30b9",
]

# Request defaults
HEADERS = {
    "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/121.0 Safari/537.36"
}

# Similarity thresholds
SIMILARITY_THRESHOLD = 0.9
SHEET2_SIMILARITY_THRESHOLD = 0.9

# LLM defaults
USE_LLM = True
LLM_ONLY = False
RESUME_LLM = False
PROCESS_LLM_SKIPPED = False
LLM_ENDPOINT = os.environ.get("LLM_ENDPOINT", "http://127.0.0.1:1234/v1/chat/completions")
LLM_MODEL = os.environ.get("LLM_MODEL", "qwen/qwen3-vl-8b")
LLM_TIMEOUT = 120
LLM_SAVE_INTERVAL = 0
LLM_CACHE = {}
LLM_IMAGE_CACHE = {}
LLM_ERROR_LOGGED = False
LLM_IMAGE_INPUT = True
LLM_IMAGE_TIMEOUT = 12
LLM_IMAGE_MAX_BYTES = 3_000_000
LLM_IMAGE_MAX_SIZE = 768
LLM_IMAGE_FORMAT = "JPEG"

# Summary limits
SUMMARY_TITLE_LIMIT = 50
SUMMARY_CONTENT_LIMIT = 150
SUMMARY_HTML_CHARS = 8000

# Feature flags
ENABLE_GOOGLE_NEWS = True
USE_PLAYWRIGHT = True
FETCH_MISSING_IMAGES = True
ENRICH_ONLY = False
ENRICH_EXISTING = False
PROGRESS_EVERY = 5
IMAGE_FETCH_TIMEOUT_SECOND = 12
IMAGE_FETCH_WORKERS = 12
OUTPUT_PAPERS_SHEET2 = False

# NewsAPI
NEWSAPI_ENDPOINT = "https://newsapi.org/v2/everything"

# PubMed image marker
PUBMED_META_IMAGE = "pubmed.ncbi.nlm.nih.gov"

# Only papers RSS mode (skip other sources)
ONLY_PAPERS_RSS = os.environ.get("ONLY_PAPERS_RSS", "").lower() in ("1", "true", "yes")

# API keys
NEWSAPI_KEY = os.environ.get("NEWSAPI_KEY", "")

# Columns to drop from output files
DROP_OUTPUT_COLUMNS = [
    "\u8003\u5bdf",
    "\u4f01\u753b\u30a2\u30a4\u30c7\u30a2",
]

# URL check timeout (seconds)
URL_CHECK_TIMEOUT = 5

try:
    from PIL import Image
    PIL_AVAILABLE = True
except ImportError:
    PIL_AVAILABLE = False

# DuckDuckGo Search
try:
    from duckduckgo_search import DDGS
    DDGS_AVAILABLE = True
except ImportError:
    DDGS_AVAILABLE = False
    print("注意: duckduckgo-search がインストールされていません。pip install duckduckgo-search を実行してください。")

try:
    from deep_translator import GoogleTranslator
    TRANSLATOR_AVAILABLE = True
except ImportError:
    TRANSLATOR_AVAILABLE = False
    print("注意: deep-translator がインストールされていません。pip install deep-translator を実行してください。")

try:
    from playwright.sync_api import sync_playwright
    PLAYWRIGHT_AVAILABLE = True
except ImportError:
    PLAYWRIGHT_AVAILABLE = False
    print("注意: playwright がインストールされていません。pip install playwright で追加できます。")

# --- 設定 ---
def load_json_file(path):
    try:
        if path and Path(path).exists():
            return json.loads(Path(path).read_text(encoding="utf-8"))
    except Exception:
        return {}
    return {}

def load_api_keys(path):
    data = load_json_file(path)
    if isinstance(data, dict):
        return data.get("newsapi_key", "") or ""
    return ""

NEWSAPI_KEY = load_api_keys(API_KEYS_PATH) or NEWSAPI_KEY

def apply_department_settings(dept, settings_path):
    """Override RSS_FEEDS/COUNTRY_SETTINGS/INTERIOR_KEYWORDS by department settings."""
    if not dept:
        return
    data = load_json_file(settings_path)
    if not data:
        return
    dept_cfg = data.get(dept) if isinstance(data, dict) else None
    if not isinstance(dept_cfg, dict):
        return
    global RSS_FEEDS, COUNTRY_SETTINGS, INTERIOR_KEYWORDS, INTERIOR_KEYWORDS_LOWER
    if isinstance(dept_cfg.get("rss_feeds"), list):
        RSS_FEEDS = dept_cfg["rss_feeds"]
    if isinstance(dept_cfg.get("country_settings"), dict):
        COUNTRY_SETTINGS = dept_cfg["country_settings"]
    if isinstance(dept_cfg.get("keywords"), list):
        INTERIOR_KEYWORDS = dept_cfg["keywords"]
        INTERIOR_KEYWORDS_LOWER = [kw.lower() for kw in INTERIOR_KEYWORDS]
    if isinstance(dept_cfg.get("synonym_groups"), list):
        global SYNONYM_GROUPS
        SYNONYM_GROUPS = dept_cfg["synonym_groups"]
    if isinstance(dept_cfg.get("subject_name"), str) and dept_cfg.get("subject_name"):
        global SUBJECT_NAME
        SUBJECT_NAME = dept_cfg.get("subject_name")
    if isinstance(dept_cfg.get("photo_target_label"), str) and dept_cfg.get("photo_target_label"):
        global PHOTO_TARGET_LABEL
        PHOTO_TARGET_LABEL = dept_cfg.get("photo_target_label")
    if isinstance(dept_cfg.get("prompt_path"), str) and dept_cfg.get("prompt_path"):
        global PROMPT_PATH, PROMPT_TEXT
        PROMPT_PATH = Path(dept_cfg.get("prompt_path"))
        PROMPT_TEXT = ""


def get_domain(url):
    try:
        return urlparse(url).netloc
    except:
        return ""

def is_similar(text1, text2):
    if not text1 or not text2:
        return False
    ratio = difflib.SequenceMatcher(None, str(text1).lower(), str(text2).lower()).ratio()
    return ratio > SIMILARITY_THRESHOLD

def is_interior_related(title, content=""):
    text = (str(title) + " " + str(content)).lower()
    return any(kw.lower() in text for kw in INTERIOR_KEYWORDS)

def parse_date(date_str):
    if not date_str:
        return None
    try:
        dt = parser.parse(str(date_str))
        return dt.strftime('%Y-%m-%d')
    except:
        return None

def extract_image_from_rss(entry):
    """RSS/Atomのエントリから画像URLを抽出"""
    try:
        media = entry.get("media_content", [])
        if media and isinstance(media, list):
            for m in media:
                url = m.get("url") or m.get("href")
                if url:
                    return url
        thumbs = entry.get("media_thumbnail", [])
        if thumbs and isinstance(thumbs, list):
            for m in thumbs:
                url = m.get("url") or m.get("href")
                if url:
                    return url
        enclosures = entry.get("enclosures", [])
        if enclosures:
            for enc in enclosures:
                url = enc.get("url")
                if url and "image" in str(enc.get("type", "")).lower():
                    return url
        links = entry.get("links", [])
        if links:
            for lnk in links:
                if lnk.get("rel") == "enclosure" and "image" in str(lnk.get("type", "")).lower():
                    url = lnk.get("href") or lnk.get("url")
                    if url:
                        return url
    except Exception:
        pass
    return ""

def is_target_date(pub_date, target_dates):
    if not target_dates:
        return True
    return pub_date in target_dates

def normalize_image_url(candidate, base_url):
    if not candidate:
        return ""
    if candidate.startswith("//"):
        candidate = "https:" + candidate
    elif candidate.startswith("/"):
        candidate = urljoin(base_url, candidate)
    if candidate.startswith("http"):
        return candidate
    return ""

def rank_image_url(url: str) -> int:
    """Prefer higher-quality Yahoo (yimg.jp) image variants when multiple candidates exist."""
    if not url:
        return -999
    score = 0
    lower = url.lower()
    if "yimg.jp" in lower:
        score += 10
        # Prefer URLs with explicit size/quality params
        if "pri=" in lower:
            score += 5
        if "w=" in lower or "h=" in lower:
            score += 3
        # Prefer -000-2- over -000-1- when both exist
        if re.search(r"-000-2-", lower):
            score += 4
        if re.search(r"-000-1-", lower):
            score -= 1
    return score

def collect_image_candidates(soup, base_url):
    candidates = []
    metas = [
        ("property", "og:image:secure_url"),
        ("property", "og:image"),
        ("name", "og:image"),
        ("name", "twitter:image"),
        ("property", "twitter:image"),
        ("name", "twitter:image:src"),
        ("property", "twitter:image:src"),
        ("name", "citation_image"),
    ]
    for attr, val in metas:
        m = soup.find("meta", attrs={attr: val})
        if m and m.get("content"):
            candidates.append(m.get("content"))
    link_img = soup.find("link", rel="image_src")
    if link_img and link_img.get("href"):
        candidates.append(link_img.get("href"))
    img = soup.find("img", src=True)
    if img:
        candidates.append(img.get("src"))
    img_ds = soup.find("img", attrs={"data-src": True})
    if img_ds:
        candidates.append(img_ds.get("data-src"))
    img_dsset = soup.find("img", attrs={"data-srcset": True})
    if img_dsset:
        srcset = img_dsset.get("data-srcset", "")
        if srcset:
            first = srcset.split(",")[0].strip().split(" ")[0]
            candidates.append(first)
    # JSON-LD image candidates
    for script in soup.find_all("script", type="application/ld+json"):
        try:
            data = json.loads(script.string)
        except Exception:
            continue
        def extract_from_json(obj):
            if isinstance(obj, str):
                return [obj]
            if isinstance(obj, list):
                out = []
                for v in obj:
                    out.extend(extract_from_json(v))
                return out
            if isinstance(obj, dict):
                out = []
                for key in ["image", "thumbnail", "logo"]:
                    if key in obj:
                        out.extend(extract_from_json(obj[key]))
                return out
            return []
        candidates.extend(extract_from_json(data))
    # MSN/Bing: pick direct img-s-msn URL
    text = soup.decode()
    match = re.search(r'https?://img-[\w.-]+/[^"\'\s>]+\.img[^"\'\s>]*', text)
    if match:
        candidates.append(match.group(0))
    normalized = []
    for cand in candidates:
        norm = normalize_image_url(cand, base_url)
        if norm:
            normalized.append(norm)
    if normalized:
        normalized.sort(key=rank_image_url, reverse=True)
    return normalized

def is_pubmed_or_pmc_url(url):
    if not url:
        return False
    lower = url.lower()
    host = urlparse(lower).netloc
    return (
        host.endswith("pubmed.ncbi.nlm.nih.gov")
        or host.endswith("pmc.ncbi.nlm.nih.gov")
        or "/pmc/" in lower
    )

def is_pubmed_placeholder_image(url):
    if not url:
        return False
    return PUBMED_META_IMAGE in url

def extract_pmc_url_from_html(html):
    if not html:
        return ""
    patterns = [
        r"https?://pmc\.ncbi\.nlm\.nih\.gov/articles/PMC\d+/?",
        r"https?://www\.ncbi\.nlm\.nih\.gov/pmc/articles/PMC\d+/?",
    ]
    for pat in patterns:
        m = re.search(pat, html)
        if m:
            return m.group(0)
    m = re.search(r"/pmc/articles/PMC\d+/?", html)
    if m:
        return urljoin("https://www.ncbi.nlm.nih.gov", m.group(0))
    return ""

def extract_pmc_figure_image(soup, base_url):
    for img in soup.select("figure img[src], img[src]"):
        src = img.get("src")
        if not src:
            continue
        lower = src.lower()
        if any(x in lower for x in ["logo", "icon", "sprite", "blank", "spacer", "placeholder"]):
            continue
        if "pmc/articles" in lower or "/bin/" in lower or "/pmc/" in lower:
            norm = normalize_image_url(src, base_url)
            if norm:
                return norm
    return ""

def fetch_image_from_page(url, timeout=10):
    """Fetch og:image or figure image from the article page."""
    if not url:
        return ""
    try:
        resp = requests.get(url, headers=HEADERS, timeout=timeout)
        if resp.status_code != 200:
            return ""
        html = resp.text
        soup = BeautifulSoup(html, "html.parser")
        candidates = collect_image_candidates(soup, url)
        image_url = candidates[0] if candidates else ""

        if is_pubmed_or_pmc_url(url) or is_pubmed_placeholder_image(image_url):
            pmc_url = extract_pmc_url_from_html(html)
            if pmc_url and pmc_url != url:
                pmc_img = fetch_image_from_page(pmc_url, timeout=timeout)
                if pmc_img and not is_pubmed_placeholder_image(pmc_img):
                    return pmc_img
            pmc_img_inline = extract_pmc_figure_image(soup, url)
            if pmc_img_inline:
                return pmc_img_inline

        return image_url
    except Exception:
        return ""

def resolve_final_url(url, timeout=5):
    """リダイレクトを解決して最終URLを返す（失敗時は元URL）"""
    if not url:
        return url
    try:
        parsed = urlparse(url)
        # Google NewsのRSSリンクは ?url= や中間ページ経由なので優先的に実URLを抽出
        qs = parse_qs(parsed.query)
        if parsed.netloc.endswith("news.google.com"):
            if "url" in qs and qs["url"]:
                return qs["url"][0]
            # base64相当のペイロードからURLを抽出
            try:
                payload = parsed.path.split("/")[-1].split("?")[0]
                pad = "=" * (-len(payload) % 4)
                decoded = base64.urlsafe_b64decode(payload + pad).decode("utf-8", errors="ignore")
                m = re.search(r"https?://[^\s'\"]+", decoded)
                if m:
                    return m.group(0)
            except Exception:
                pass
        resp = requests.get(url, headers=HEADERS, timeout=timeout, allow_redirects=True)
        final_url = resp.url or url
        # news.google.com の記事ページから canonical を拾う
        if urlparse(final_url).netloc.endswith("news.google.com"):
            try:
                soup = BeautifulSoup(resp.text, "html.parser")
                canonical = soup.find("link", rel="canonical")
                if canonical and canonical.get("href"):
                    final_url = canonical.get("href")
                else:
                    og = soup.find("meta", property="og:url")
                    if og and og.get("content"):
                        final_url = og.get("content")
            except Exception:
                pass
        return final_url
    except Exception:
        return url

def resolve_with_playwright(url, timeout_ms=15000):
    """PlaywrightでJSレンダリング後のURLと画像を取得（Playwrightがある場合のみ）"""
    if not PLAYWRIGHT_AVAILABLE or not USE_PLAYWRIGHT:
        return None, None
    try:
        with sync_playwright() as p:
            browser = p.chromium.launch(headless=True)
            page = browser.new_page()
            page.goto(url, wait_until="load", timeout=timeout_ms)
            final_url = page.url
            image_url = ""
            # canonical
            canonical = page.locator("link[rel=canonical]").first
            if canonical.count() > 0:
                href = canonical.get_attribute("href")
                if href:
                    final_url = href
            # og:image
            og_image = page.locator("meta[property='og:image'], meta[name='og:image'], meta[name='twitter:image'], meta[property='twitter:image']").first
            if og_image.count() > 0:
                content = og_image.get_attribute("content")
                if content:
                    image_url = content
            # ページ中の外部リンクをサーチ（非google）
            if "news.google.com" in (final_url or ""):
                html = page.content()
                links = re.findall(r"https?://[^\"'\\s<>]+", html)
                for l in links:
                    if ("google" not in l) and ("gstatic" not in l) and ("youtube" not in l):
                        final_url = l
                        break
            browser.close()
            return final_url, image_url
    except Exception:
        return None, None

def search_article_url(title, country=""):
    """記事タイトルでWeb検索し、最上位のURLを返す（DDGSが使える場合のみ）"""
    if not DDGS_AVAILABLE:
        return ""
    query = title
    try:
        with DDGS() as ddgs:
            results = ddgs.text(query, region=None, safesearch="off", max_results=3)
            for r in results:
                url = r.get("href") or r.get("url")
                if url and url.startswith("http"):
                    return url
    except Exception:
        return ""
    return ""

def load_prompt_text():
    if not PROMPT_PATH.exists():
        return ""
    for enc in ("utf-8", "utf-8-sig", "cp932"):
        try:
            text = PROMPT_PATH.read_text(encoding=enc)
            if text and text.strip():
                return text
        except Exception:
            continue
    return ""

def image_url_to_data_url(image_url):
    """画像URLをLLM入力向けのdata URLに変換（サイズ縮小）"""
    if not image_url or not PIL_AVAILABLE:
        return ""
    if image_url in LLM_IMAGE_CACHE:
        return LLM_IMAGE_CACHE[image_url]
    try:
        resp = requests.get(image_url, headers=HEADERS, timeout=LLM_IMAGE_TIMEOUT)
        if resp.status_code >= 400:
            return ""
        clen = resp.headers.get("Content-Length")
        if clen:
            try:
                if int(clen) > LLM_IMAGE_MAX_BYTES:
                    return ""
            except Exception:
                pass
        data = resp.content
        if not data:
            return ""
        img = Image.open(BytesIO(data))
        img = img.convert("RGB")
        img.thumbnail((LLM_IMAGE_MAX_SIZE, LLM_IMAGE_MAX_SIZE), Image.LANCZOS)
        buf = BytesIO()
        fmt = (LLM_IMAGE_FORMAT or "JPEG").upper()
        img.save(buf, format=fmt, quality=80, optimize=True)
        mime = "jpeg" if fmt in ("JPG", "JPEG") else fmt.lower()
        b64 = base64.b64encode(buf.getvalue()).decode("ascii")
        data_url = f"data:image/{mime};base64,{b64}"
        LLM_IMAGE_CACHE[image_url] = data_url
        return data_url
    except Exception:
        return ""

def call_llm_classify(title, content, image_url="", mode="both"):
    """ローカルLLMで対象判定と画像判定を実施する"""
    global PROMPT_TEXT, LLM_ERROR_LOGGED
    if not USE_LLM:
        return "", ""
    if not PROMPT_TEXT:
        PROMPT_TEXT = load_prompt_text()
    if not PROMPT_TEXT:
        if not LLM_ERROR_LOGGED:
            print("  ✗ プロンプト.md を読み込めませんでした（エンコード/パス確認）")
            LLM_ERROR_LOGGED = True
        return "", ""
    cache_key = (title, content, image_url, mode)
    if cache_key in LLM_CACHE:
        return LLM_CACHE[cache_key]
    if not USE_LLM:
        return "", ""
    if mode == "relevance":
        prompt = (
            f"{PROMPT_TEXT}\n\nタイトル:\n{title}\n\n本文:\n{content}\n\n"
            "以下を簡潔にJSONで回答してください。\n"
            "keys=['relevance']\n"
            "relevanceは対象ならtrue/false。理由は不要。"
        )
    elif mode == "photo":
        prompt = (
            f"{PROMPT_TEXT}\n\nタイトル:\n{title}\n\n本文:\n{content}\n\n"
            "以下を簡潔にJSONで回答してください。\n"
            "keys=['has_target_photo']\n"
            f"has_target_photoは「{PHOTO_TARGET_LABEL}」の写真が本文や画像から確認できそうならtrue/false。理由は不要。"
        )
    else:
        prompt = (
            f"{PROMPT_TEXT}\n\nタイトル:\n{title}\n\n本文:\n{content}\n\n"
            "以下を簡潔にJSONで回答してください。\n"
            "keys=['relevance','has_target_photo']\n"
            "relevanceは対象ならtrue/false。\n"
            f"has_target_photoは「{PHOTO_TARGET_LABEL}」の写真が本文や画像から確認できそうならtrue/false。理由は不要。"
        )
    try:
        content_payload = prompt
        if mode != "relevance" and LLM_IMAGE_INPUT and isinstance(image_url, str) and image_url.startswith("http"):
            data_url = image_url_to_data_url(image_url)
            if data_url:
                content_payload = [
                    {"type": "text", "text": prompt},
                    {"type": "image_url", "image_url": {"url": data_url}},
                ]
        payload = {
            "model": LLM_MODEL,
            "messages": [
                {"role": "user", "content": content_payload}
            ],
            "temperature": 0.2,
        }
        resp = requests.post(LLM_ENDPOINT, json=payload, timeout=LLM_TIMEOUT)
        if resp.status_code != 200:
            if isinstance(content_payload, list):
                payload["messages"][0]["content"] = prompt
                resp = requests.post(LLM_ENDPOINT, json=payload, timeout=LLM_TIMEOUT)
            if resp.status_code != 200:
                if not LLM_ERROR_LOGGED:
                    print(f"  ✗ LLM呼び出しエラー: HTTP {resp.status_code}")
                    print(resp.text[:200])
                    LLM_ERROR_LOGGED = True
                return "", ""
        txt = resp.json().get("choices", [{}])[0].get("message", {}).get("content", "")
        rel = ""
        photo = ""
        try:
            if mode in ("both", "relevance"):
                m = re.search(r'"relevance"\s*:\s*(true|false)', txt, re.IGNORECASE)
                if m:
                    rel = "対象" if m.group(1).lower() == "true" else "非対象"
            if mode in ("both", "photo"):
                m2 = re.search(r'"has_target_photo"\s*:\s*(true|false)', txt, re.IGNORECASE)
                if m2:
                    photo = "あり" if m2.group(1).lower() == "true" else "なし"
        except Exception:
            pass
        result = (rel, photo)
        LLM_CACHE[cache_key] = result
        return result
    except Exception as e:
        if not LLM_ERROR_LOGGED:
            print(f"  LLM error: {e}")
            LLM_ERROR_LOGGED = True
        return "", ""

def check_url_ok(url, is_image=False, timeout=URL_CHECK_TIMEOUT):
    """URLが有効ならTrue、無効ならFalse"""
    if not url or not isinstance(url, str):
        return False
    lower = url.lower()
    is_yimg = "yimg.jp" in lower
    if "favicon" in lower or "sprite" in lower:
        return False
    if is_image and ("logo" in lower or "placeholder" in lower or "googleusercontent" in lower):
        return False
    try:
        resp = requests.head(url, headers=HEADERS, timeout=timeout, allow_redirects=True)
        if resp.status_code >= 400:
            if is_image and is_yimg:
                try:
                    get_resp = requests.get(url, headers=HEADERS, timeout=timeout, stream=True)
                    if get_resp.status_code >= 400:
                        return False
                    ctype = get_resp.headers.get("Content-Type", "").lower()
                    return "image" in ctype
                except Exception:
                    return False
            return False
        ctype = resp.headers.get("Content-Type", "").lower()
        clen = resp.headers.get("Content-Length")
        if is_image:
            if "image" not in ctype:
                if is_yimg:
                    try:
                        get_resp = requests.get(url, headers=HEADERS, timeout=timeout, stream=True)
                        if get_resp.status_code >= 400:
                            return False
                        ctype = get_resp.headers.get("Content-Type", "").lower()
                        if "image" not in ctype:
                            return False
                    except Exception:
                        return False
                else:
                    return False
            try:
                if clen and int(clen) < 1500:
                    return False
            except Exception:
                pass
        return True
    except Exception:
        if is_image and is_yimg:
            try:
                get_resp = requests.get(url, headers=HEADERS, timeout=timeout, stream=True)
                if get_resp.status_code >= 400:
                    return False
                ctype = get_resp.headers.get("Content-Type", "").lower()
                return "image" in ctype
            except Exception:
                return False
        return False

def bulk_fetch_images(urls):
    """URLリストを並列取得して画像URLを返す辞書を作成"""
    unique = []
    seen = set()
    for u in urls:
        if not u or u in seen:
            continue
        unique.append(u)
        seen.add(u)
    results = {u: "" for u in unique}
    if not unique:
        return results
    with ThreadPoolExecutor(max_workers=IMAGE_FETCH_WORKERS) as executor:
        future_to_url = {executor.submit(fetch_image_from_page, u): u for u in unique}
        for fut in as_completed(future_to_url):
            u = future_to_url[fut]
            try:
                results[u] = fut.result()
            except Exception:
                results[u] = ""
    # 二次: 失敗分を長めタイムアウトで再取得
    secondary_targets = [u for u, img in results.items() if is_missing_url(img)]
    if secondary_targets:
        with ThreadPoolExecutor(max_workers=max(4, IMAGE_FETCH_WORKERS // 2)) as executor:
            future_to_url = {executor.submit(fetch_image_from_page, u, IMAGE_FETCH_TIMEOUT_SECOND): u for u in secondary_targets}
            for fut in as_completed(future_to_url):
                u = future_to_url[fut]
                try:
                    img = fut.result()
                    if img:
                        results[u] = img
                except Exception:
                    pass
    # 三次: サイトスクリーンショットサービスでサムネ生成
    # 無効ならスキップ（空欄のまま残す）
    return results

def is_missing_url(val):
    if val is None:
        return True
    if isinstance(val, float) and math.isnan(val):
        return True
    if isinstance(val, str):
        s = val.strip()
        if s == "":
            return True
        lower = s.lower()
        if "google.com/s2/favicons" in lower:
            return True
        if lower.endswith("favicon.ico"):
            return True
        if PUBMED_META_IMAGE in lower:
            return True
    return False

def is_yimg_placeholder(url, timeout=5):
    """Detect Yahoo yimg.jp placeholder (often 404 image/gif small file)."""
    if not url or "yimg.jp" not in str(url).lower():
        return False
    try:
        resp = requests.get(url, headers=HEADERS, timeout=timeout, stream=True)
        if resp.status_code >= 400:
            return True
        ctype = (resp.headers.get("Content-Type", "") or "").lower()
        clen = resp.headers.get("Content-Length")
        if clen:
            try:
                if int(clen) < 2000:
                    return True
            except Exception:
                pass
        if "image/gif" in ctype:
            # Yahoo 404 placeholder tends to be a tiny gif
            chunk = resp.raw.read(2048, decode_content=True)
            if chunk and len(chunk) < 2000:
                return True
        return False
    except Exception:
        return True

_translation_cache = {}
_summary_cache = {}
_article_text_cache = {}

def compute_relevance(title, content=""):
    """タイトル/本文から対象分野の関連度をスコアリングする"""
    text = (str(title) + " " + str(content)).lower()
    hits = []
    seen_canon = []
    for kw in INTERIOR_KEYWORDS_LOWER:
        if kw and kw in text:
            # 同義語グループをまとめてカウント
            canon = kw
            for group in SYNONYM_GROUPS:
                if kw in group:
                    canon = sorted(group)[0]
                    break
            if canon not in seen_canon:
                seen_canon.append(canon)
            hits.append(kw)
    unique_hits = seen_canon
    score = min(1.0, len(unique_hits) / 5.0)  # 5種以上で1.0
    if score >= 0.7:
        label = "高"
    elif score >= 0.4:
        label = "中"
    else:
        label = "低"
    return round(score, 2), label, unique_hits

def normalize_text(text):
    return re.sub(r"\s+", " ", str(text or "")).strip()

def has_japanese_kana(text):
    return re.search(r"[ぁ-ゟァ-ヿ]", str(text or "")) is not None

def has_kanji(text):
    return re.search(r"[\u4e00-\u9fff]", str(text or "")) is not None

def kana_kanji_counts(text):
    s = str(text or "")
    kana = len(re.findall(r"[ぁ-ゟァ-ヿ]", s))
    kanji = len(re.findall(r"[\u4e00-\u9fff]", s))
    return kana, kanji

def is_kana_heavy(text, ratio=0.65, min_kana=8, kana_over_kanji=1.5):
    kana, kanji = kana_kanji_counts(text)
    total = kana + kanji
    if total == 0 or kana < min_kana:
        return False
    if kanji == 0:
        return True
    if (kana / max(kanji, 1)) >= kana_over_kanji:
        return True
    return (kana / total) >= ratio

def contains_yen(text):
    s = str(text or "")
    tokens = ["\u5186", "\u4e07\u5186", "\u5186\u53f0", "\uffe5", "JPY", "\u00a5"]
    return any(tok in s for tok in tokens)

def extract_cjk_keywords(text, limit=8):
    s = str(text or "")
    tokens = re.findall(r"[\u4e00-\u9fff]{2,}", s)
    if not tokens:
        return []
    counts = Counter(tokens)
    return [t for t, _ in counts.most_common(limit)]

def summary_matches_source(summary, source):
    keys = extract_cjk_keywords(source)
    if not keys:
        return True
    return any(k in summary for k in keys)

def enforce_kanji_text(text):
    if not text or not USE_LLM:
        return ""
    prompt = (
        "Rewrite the following Japanese into natural Japanese with kanji.\n"
        "Avoid kana-only output and avoid phonetic kana spellings for proper nouns.\n"
        'Output JSON only: {"text":"..."}\n\n'
        f"{text}"
    )
    output = call_llm_text(prompt)
    fixed = parse_json_field(output, "text")
    if not fixed:
        fixed = normalize_text(output)
    if fixed and has_kanji(fixed):
        return fixed
    return ""

def ensure_japanese(text, fallback="", force=False, require_kanji=False):
    if not text:
        return ""
    if not force and has_japanese_kana(text):
        return text
    translated = translate_text(text, target_lang="ja", force_japanese=True, require_kanji=require_kanji)
    if translated:
        needs_kanji = require_kanji and (not has_kanji(translated) or is_kana_heavy(translated))
        if needs_kanji:
            converted = enforce_kanji_text(translated)
            if converted:
                translated = converted
            else:
                translated = ""
        if translated and (has_japanese_kana(translated) or has_kanji(translated)):
            return translated
    if fallback and fallback != text:
        translated_alt = translate_text(fallback, target_lang="ja", force_japanese=True, require_kanji=require_kanji)
        if translated_alt:
            needs_kanji = require_kanji and (not has_kanji(translated_alt) or is_kana_heavy(translated_alt))
            if needs_kanji:
                converted = enforce_kanji_text(translated_alt)
                if converted:
                    translated_alt = converted
                else:
                    translated_alt = ""
            if translated_alt:
                return translated_alt
    return translated or text

def ends_with_sentence(text):
    s = normalize_text(text)
    if not s:
        return False
    last = s[-1]
    if last in "。！？.!?":
        return True
    if last in "」』）】］" and len(s) >= 2 and s[-2] in "。！？.!?":
        return True
    return False

def trim_to_sentence(text, limit):
    s = normalize_text(text)
    if len(s) <= limit:
        return s
    boundaries = ["。", "！", "？", ".", "!", "?"]
    best = ""
    for b in boundaries:
        idx = s.rfind(b, 0, limit + 1)
        if idx != -1:
            cand = s[: idx + 1].strip()
            if len(cand) > len(best):
                best = cand
    if best:
        return best
    return s[:limit].strip()

def call_llm_text(prompt):
    payload = {
        "model": LLM_MODEL,
        "messages": [
            {"role": "user", "content": prompt}
        ],
        "temperature": 0.2,
    }
    try:
        resp = requests.post(LLM_ENDPOINT, json=payload, timeout=LLM_TIMEOUT)
        if resp.status_code != 200:
            return ""
        return resp.json().get("choices", [{}])[0].get("message", {}).get("content", "")
    except Exception:
        return ""

def parse_json_field(text, key):
    try:
        data = json.loads(text)
        return normalize_text(data.get(key, ""))
    except Exception:
        m = re.search(rf'"{re.escape(key)}"\s*:\s*"(.*?)"', text)
        if m:
            return normalize_text(m.group(1))
    return ""

def translate_text(text, target_lang="ja", force_japanese=False, require_kanji=False):
    """翻訳（ローカルLLM）。失敗時は空文字を返す"""
    if not text:
        return ""
    key = (text, target_lang, force_japanese, require_kanji)
    if key in _translation_cache:
        return _translation_cache[key]
    if not USE_LLM:
        return ""
    if target_lang == "ja" and force_japanese:
        prompt = (
            "次の文章を自然な日本語に翻訳してください。必ず日本語で、漢字とひらがなを適切に混ぜてください。"
            "ひらがな/カタカナだけの出力は禁止です。"
            '出力はJSONのみ。形式: {"translation":"..."}\\n\\n'
            f"{text}"
        )
    else:
        prompt = (
            f"次の文章を{target_lang}に翻訳してください。出力はJSONのみ。"
            f'形式: {{"translation":"..."}}\\n\\n{text}'
        )
    if target_lang == "ja" and force_japanese:
        # 再定義: 固有名詞の当て字を抑制
        prompt = (
            "次の文章を自然な日本語に翻訳してください。必ず日本語で、漢字とひらがなを適切に混ぜてください。\n"
            "固有名詞は英字または一般的な表記を使い、ひらがな主体の当て字は避けてください。\n"
            "ひらがな/カタカナだけの出力は禁止です。\n"
            '出力はJSONのみ。形式: {"translation":"..."}\n\n'
            f"{text}"
        )
    output = call_llm_text(prompt)
    translated = parse_json_field(output, "translation")
    if target_lang == "ja" and force_japanese and require_kanji and translated and not has_kanji(translated):
        strict_prompt = (
            "次の文章を日本語に翻訳してください。必ず漢字を含め、自然な日本語にしてください。"
            "ひらがな/カタカナだけの出力は禁止です。"
            '出力はJSONのみ。形式: {"translation":"..."}\\n\\n'
            f"{text}"
        )
        # 再定義: 固有名詞の当て字を抑制
        strict_prompt = (
            "次の文章を日本語に翻訳してください。必ず漢字を含め、自然な日本語にしてください。\n"
            "固有名詞は英字または一般的な表記を使い、ひらがな主体の当て字は避けてください。\n"
            "ひらがな/カタカナだけの出力は禁止です。\n"
            '出力はJSONのみ。形式: {"translation":"..."}\n\n'
            f"{text}"
        )
        strict_prompt = strict_prompt + "\nCurrency rule: Keep original currency/units; do not convert to JPY or invent amounts."
        strict_output = call_llm_text(strict_prompt)
        strict_translated = parse_json_field(strict_output, "translation")
        if strict_translated:
            translated = strict_translated
    if target_lang == "ja" and not force_japanese and translated and not has_japanese_kana(translated):
        strict_prompt = (
            "次の文章を日本語に翻訳してください。必ず日本語で、漢字とひらがなを適切に混ぜてください。"
            "ひらがな/カタカナだけの出力は禁止です。"
            '出力はJSONのみ。形式: {"translation":"..."}\\n\\n'
            f"{text}"
        )
        strict_output = call_llm_text(strict_prompt)
        strict_translated = parse_json_field(strict_output, "translation")
        if strict_translated:
            translated = strict_translated
    if translated:
        _translation_cache[key] = translated
    return translated

def fetch_article_text(url):
    if not url or not isinstance(url, str) or not url.startswith("http"):
        return ""
    if url in _article_text_cache:
        return _article_text_cache[url]
    try:
        resp = requests.get(url, headers=HEADERS, timeout=15)
        if resp.status_code != 200:
            return ""
        html = resp.text
    except Exception:
        return ""
    soup = BeautifulSoup(html, "html.parser")
    for tag in soup(["script", "style", "noscript"]):
        tag.decompose()
    main = soup.find("article") or soup.find("main") or soup.body or soup
    text = " ".join(main.stripped_strings)
    text = normalize_text(text)
    if len(text) > SUMMARY_HTML_CHARS:
        text = text[:SUMMARY_HTML_CHARS]
    _article_text_cache[url] = text
    return text

def summarize_article(title, content, url, country=""):
    """URL本文 + 既存タイトル/本文から日本語要約を生成"""
    cache_key = (url or "", title or "", content or "")
    if cache_key in _summary_cache:
        return _summary_cache[cache_key]
    if not USE_LLM:
        fallback_title = normalize_text(title)
        fallback_body = normalize_text(content)
        fallback_title = trim_to_sentence(fallback_title, SUMMARY_TITLE_LIMIT)
        fallback_body = trim_to_sentence(fallback_body, SUMMARY_CONTENT_LIMIT)
        if fallback_body and not ends_with_sentence(fallback_body):
            fallback_body = fallback_body + "。"
        _summary_cache[cache_key] = (fallback_title, fallback_body)
        return fallback_title, fallback_body

    html_text = ""
    if isinstance(url, str) and url:
        if "gasgoo.com" not in url:
            html_text = fetch_article_text(url)
    source_text = f"{title} {content}"
    cn_keywords = extract_cjk_keywords(source_text) if country == "中国" else []
    keep_terms = ""
    if cn_keywords:
        keep_terms = "Use these terms as-is (do not translate): " + ", ".join(cn_keywords[:6]) + "\\n"
    currency_rule = "Currency rule: Keep original currency/units; do not convert to JPY or invent amounts.\\n"
    prompt = (
        "以下の情報を統合して、日本語で要約してください。\\n"
        f"条件1: タイトルは{SUMMARY_TITLE_LIMIT}字以内（文を途中で切らない）。\\n"
        f"条件2: 内容は{SUMMARY_CONTENT_LIMIT}字以内（文を途中で切らない）。\\n"
        "条件3: 事実ベースで簡潔に。\\n"
        "条件4: 可能なら固有名詞/数値/企業名など本文の具体情報を1つ以上含める。\\n"
        "条件5: 内容は必ず句点で終える。\\n"
        "条件6: 出力はJSONのみ。形式: {\"title\":\"...\",\"summary\":\"...\"}\\n\\n"
        f"既存タイトル: {title}\\n"
        f"既存内容: {content}\\n"
        f"本文HTML抽出: {html_text}\\n"
        f"URL: {url}"
    )
    if country == "中国":
        prompt = (
            "以下の情報を統合して、日本語で要約してください。\n"
            f"条件1: タイトルは{SUMMARY_TITLE_LIMIT}字以内、文を途中で切らない。\n"
            f"条件2: 内容は{SUMMARY_CONTENT_LIMIT}字以内、文を途中で切らない。\n"
            "条件3: 事実ベースで簡潔に。\n"
            "条件4: 可能なら固有名詞/数値/企業名など具体情報を2つ以上含める。\n"
            "条件5: 内容は必ず句点で終える。\n"
            "条件6: 中国記事は漢字や一般的なカタカナ表記を優先し、ひらがな主体の当て字は使わない。\n"
            "条件7: 固有名詞は英字表記か一般的な日本語表記（例: Huawei=ファーウェイ、Dongfeng=東風）を使う。\n"
            "条件8: 不明な固有名詞は原文の中国語表記を維持する。\n"
            "条件9: 出力はJSONのみ。形式: {\"title\":\"...\",\"summary\":\"...\"}\n\n"
            f"既存タイトル: {title}\n"
            f"既存内容: {content}\n"
            f"本文HTML抽出: {html_text}\n"
            f"URL: {url}"
        )
    if keep_terms:
        prompt = prompt + "\n" + keep_terms
    prompt = prompt + "\n" + currency_rule
    output = call_llm_text(prompt)
    summary_title = parse_json_field(output, "title")
    summary_body = parse_json_field(output, "summary")

    if len(summary_title) > SUMMARY_TITLE_LIMIT or len(summary_body) > SUMMARY_CONTENT_LIMIT or not ends_with_sentence(summary_body):
        refine_prompt = (
            f"次のJSONのtitleとsummaryを条件内に収めて書き直してください。\\n"
            f"条件1: タイトルは{SUMMARY_TITLE_LIMIT}字以内（文を途中で切らない）。\\n"
            f"条件2: 内容は{SUMMARY_CONTENT_LIMIT}字以内（文を途中で切らない）。\\n"
            "条件3: 内容は必ず句点で終える。\\n"
            "出力はJSONのみ。\\n\\n"
            f'{{"title":"{summary_title}","summary":"{summary_body}"}}'
        )
        refined = call_llm_text(refine_prompt)
        summary_title = parse_json_field(refined, "title") or summary_title
        summary_body = parse_json_field(refined, "summary") or summary_body

    if not summary_title:
        summary_title = translate_text(title) or normalize_text(title)
    if not summary_body:
        summary_body = translate_text(content) or normalize_text(content)

    force_jp = (country == "中国")
    summary_title = ensure_japanese(summary_title, title, force=force_jp, require_kanji=force_jp)
    summary_body = ensure_japanese(summary_body, content, force=force_jp, require_kanji=force_jp)

    # Retry if summary looks off-topic for CN sources
    source_text = f"{title} {content}"
    if country == "中国":
        combined = f"{summary_title} {summary_body}"
        if not summary_matches_source(combined, source_text):
            keywords = ", ".join(extract_cjk_keywords(source_text))
            retry_prompt = (
                "Summarize strictly based on the provided title/content. Avoid generic AI boilerplate.\n"
                f"Title limit: {SUMMARY_TITLE_LIMIT} chars. Summary limit: {SUMMARY_CONTENT_LIMIT} chars.\n"
                "Include at least one keyword from the list if possible.\n"
                "Output JSON only: {\"title\":\"...\",\"summary\":\"...\"}\n\n"
                f"Title: {title}\n"
                f"Content: {content}\n"
                f"Keywords: {keywords}\n"
            )
            retry = call_llm_text(retry_prompt)
            retry_title = parse_json_field(retry, "title")
            retry_body = parse_json_field(retry, "summary")
            if retry_title:
                summary_title = ensure_japanese(retry_title, title, force=force_jp, require_kanji=force_jp)
            if retry_body:
                summary_body = ensure_japanese(retry_body, content, force=force_jp, require_kanji=force_jp)

    # Avoid adding Japanese company suffixes if not present in source
    if country == "中国" and "株式会社" in f"{summary_title} {summary_body}" and "株式会社" not in source_text:
        fix_prompt = (
            "Fix this Japanese summary to avoid Japanese company suffixes like '株式会社'.\n"
            "Keep original Chinese company names as-is if unsure.\n"
            'Output JSON only: {"title":"...","summary":"..."}\n\n'
            f"Source: {source_text}\n"
            f"Title: {summary_title}\n"
            f"Summary: {summary_body}\n"
        )
        fixed = call_llm_text(fix_prompt)
        fixed_title = parse_json_field(fixed, "title")
        fixed_body = parse_json_field(fixed, "summary")
        if fixed_title:
            summary_title = ensure_japanese(fixed_title, title, force=force_jp, require_kanji=force_jp)
        if fixed_body:
            summary_body = ensure_japanese(fixed_body, content, force=force_jp, require_kanji=force_jp)

    source_text = f"{title} {content} {html_text}"
    if not contains_yen(source_text) and (contains_yen(summary_title) or contains_yen(summary_body)):
        fix_prompt = (
            "Fix currency/amounts in this Japanese summary.\n"
            "Rules: Keep original currency and units as in source; do NOT convert to JPY; do NOT invent amounts.\n"
            "If currency is unknown, remove the amount.\n"
            'Output JSON only: {"title":"...","summary":"..."}\n\n'
            f"Source: {source_text}\n"
            f"Summary title: {summary_title}\n"
            f"Summary body: {summary_body}\n"
        )
        fixed = call_llm_text(fix_prompt)
        fixed_title = parse_json_field(fixed, "title")
        fixed_body = parse_json_field(fixed, "summary")
        if fixed_title:
            summary_title = fixed_title
        if fixed_body:
            summary_body = fixed_body

    if len(summary_title) > SUMMARY_TITLE_LIMIT:
        summary_title = trim_to_sentence(summary_title, SUMMARY_TITLE_LIMIT)
    if len(summary_body) > SUMMARY_CONTENT_LIMIT:
        summary_body = trim_to_sentence(summary_body, SUMMARY_CONTENT_LIMIT)
    if summary_body and not ends_with_sentence(summary_body):
        if len(summary_body) < SUMMARY_CONTENT_LIMIT:
            summary_body = summary_body + "。"

    _summary_cache[cache_key] = (summary_title, summary_body)
    return summary_title, summary_body

def save_with_hyperlinks(df, filename):
    drop_cols = [c for c in DROP_OUTPUT_COLUMNS if c in df.columns]
    if drop_cols:
        df = df.drop(columns=drop_cols)
    """DataFrameをCSV/Excelに保存する。Excelの場合はURL列をハイパーリンク化する"""
    if str(filename).lower().endswith(".csv"):
        csv_df = df.copy()
        for col in csv_df.columns:
            if csv_df[col].dtype == object:
                csv_df[col] = csv_df[col].astype(str).str.replace("\r\n", " ", regex=False)
                csv_df[col] = csv_df[col].str.replace("\n", " ", regex=False)
                csv_df[col] = csv_df[col].str.replace("\r", " ", regex=False)
        csv_df.to_csv(filename, index=False, quoting=csv.QUOTE_ALL, encoding="utf-8", lineterminator="\n")
        return

    df.to_excel(filename, index=False)

    wb = load_workbook(filename)
    ws = wb.active

    def apply_hyperlink(column_name):
        col_idx = None
        for col in range(1, ws.max_column + 1):
            if ws.cell(row=1, column=col).value == column_name:
                col_idx = col
                break
        if not col_idx:
            return
        for row in range(2, ws.max_row + 1):
            cell = ws.cell(row=row, column=col_idx)
            url = cell.value
            if url and isinstance(url, str) and url.startswith("http"):
                cell.hyperlink = url
                cell.style = "Hyperlink"

    apply_hyperlink("URL")
    apply_hyperlink("画像URL")

    wb.save(filename)

def load_existing_data(path):
    if not path or not os.path.exists(path):
        return pd.DataFrame()
    try:
        if str(path).lower().endswith(".csv"):
            return pd.read_csv(path, encoding="utf-8", keep_default_na=False)
        return pd.read_excel(path)
    except Exception:
        return pd.DataFrame()

def build_sheet2_and_csv(df, excel_path, target_dates):
    """LLM判定対象・画像URLあり・対象日付の一覧をSheet2とCSVに出力"""
    if df is None or df.empty:
        print("  Sheet2/CSV: 対象データなし")
        return
    col_country = "国"
    col_title = "タイトル"
    col_title_jp = "タイトル（日本語）"
    col_content = "内容"
    col_content_jp = "内容（日本語）"
    col_site = "出展サイト"
    col_image = "画像URL"
    col_url = "URL"
    col_llm = "LLM判定"
    col_llm_post = "LLM後処理"
    col_date = "日付"

    work = df.copy()
    for col in [col_country, col_title, col_title_jp, col_content, col_content_jp, col_site, col_image, col_url, col_llm, col_llm_post, col_date]:
        if col not in work.columns:
            work[col] = ""

    # 日本語欄が空なら原文で補完
    work[col_title_jp] = work[col_title_jp].where(work[col_title_jp].astype(str).str.strip().ne(""), work[col_title])
    work[col_content_jp] = work[col_content_jp].where(work[col_content_jp].astype(str).str.strip().ne(""), work[col_content])

    filtered = work
    if target_dates:
        date_set = set(target_dates)
        dt = pd.to_datetime(filtered[col_date], errors="coerce")
        filtered = filtered[dt.dt.strftime("%Y-%m-%d").isin(date_set)]
    else:
        dt = pd.to_datetime(filtered[col_date], errors="coerce")
        if dt.notna().any():
            max_date = dt.max().date()
            filtered = filtered[dt.dt.date == max_date]

    # 画像URLあり（論文は例外で許容）
    filtered = filtered[(~filtered[col_image].apply(is_missing_url)) | (filtered[col_country] == "論文")]

    # 国別にLLM判定=対象を優先し、10件未満なら非対象も追加（類似は極力除外）
    filtered = filtered.copy()
    # related_sort_applied
    score_col = "\u95a2\u9023\u5ea6\u30b9\u30b3\u30a2"
    if score_col in filtered.columns:
        filtered[score_col] = pd.to_numeric(filtered[score_col], errors="coerce")
    filtered["_order"] = range(len(filtered))
    llm_flag = filtered[col_llm].astype(str).str.strip()
    result_groups = []
    for country_name, group in filtered.groupby(col_country, sort=False):
        if score_col in group.columns:
            group = group.sort_values([score_col, "_order"], ascending=[False, True])
        else:
            group = group.sort_values("_order")
        target_group = group[llm_flag.loc[group.index] == "対象"]
        extras = group[llm_flag.loc[group.index] != "対象"]
        selected = []
        selected_idx = set()
        selected_texts = []
        limit = None if country_name == "論文" else 10

        def text_key(row):
            return f"{row.get(col_title_jp, '')} {row.get(col_content_jp, '')}".strip()

        def is_similar_text(text):
            for s in selected_texts:
                if difflib.SequenceMatcher(None, text, s).ratio() >= SHEET2_SIMILARITY_THRESHOLD:
                    return True
            return False

        def add_rows(df_rows, allow_similar=False):
            nonlocal selected, selected_idx, selected_texts
            for idx, row in df_rows.iterrows():
                if idx in selected_idx:
                    continue
                text = text_key(row)
                if not allow_similar and text and is_similar_text(text):
                    continue
                selected.append(row)
                selected_idx.add(idx)
                if text:
                    selected_texts.append(text)
                if limit is not None and len(selected) >= limit:
                    return

        add_rows(target_group, allow_similar=False)
        if len(selected) < 10:
            add_rows(extras, allow_similar=False)
        if len(selected) < 10:
            add_rows(pd.concat([target_group, extras], ignore_index=False), allow_similar=True)

        if selected:
            result_groups.append(pd.DataFrame(selected))
    if result_groups:
        filtered = pd.concat(result_groups, ignore_index=False)
    else:
        filtered = filtered.iloc[0:0]
    filtered = filtered.sort_values("_order")
    filtered = filtered.drop(columns=["_order"], errors="ignore")

    # Sheet2に載るスキップ済み（日本語未補完）を補完
    if not filtered.empty:
        need_indices = []
        for idx, row in filtered.iterrows():
            title_jp = str(row.get(col_title_jp, "")).strip()
            content_jp = str(row.get(col_content_jp, "")).strip()
            llm_post = str(row.get(col_llm_post, "")).strip()
            if not title_jp or not content_jp or llm_post == "スキップ":
                need_indices.append(idx)
        if need_indices:
            print(f"  Sheet2補完: {len(need_indices)}件")
            for n, idx in enumerate(need_indices, 1):
                row = work.loc[idx]
                summary_title, summary_body = summarize_article(
                    str(row.get(col_title, "")),
                    str(row.get(col_content, "")),
                    str(row.get(col_url, "")),
                    str(row.get(col_country, "")),
                )
                if summary_title:
                    work.at[idx, col_title_jp] = summary_title
                    filtered.at[idx, col_title_jp] = summary_title
                if summary_body:
                    work.at[idx, col_content_jp] = summary_body
                    filtered.at[idx, col_content_jp] = summary_body
                if str(row.get(col_llm_post, "")).strip() == "スキップ":
                    work.at[idx, col_llm_post] = "補完"
                    filtered.at[idx, col_llm_post] = "補完"
                if n == 1 or n % PROGRESS_EVERY == 0 or n == len(need_indices):
                    print(f"    進捗: {n}/{len(need_indices)}")
            try:
                save_with_hyperlinks(work, excel_path)
            except Exception:
                pass

    sheet_cols = [col_country, col_date, col_title_jp, col_content_jp, col_site, col_image, col_url]
    sheet2_df = filtered[sheet_cols].copy()
    if col_date in sheet2_df.columns:
        sheet2_df[col_date] = pd.to_datetime(sheet2_df[col_date], errors="coerce").dt.strftime("%Y-%m-%d")

    excel_path_str = str(excel_path).lower()
    if not excel_path_str.endswith(".csv"):
        # Sheet2書き込み
        with pd.ExcelWriter(excel_path, engine="openpyxl", mode="a", if_sheet_exists="replace") as writer:
            sheet2_df.to_excel(writer, index=False, sheet_name="Sheet2")

        # Sheet2のURLをハイパーリンク化
        wb = load_workbook(excel_path)
        if "Sheet2" in wb.sheetnames:
            ws = wb["Sheet2"]
            headers_row = [cell.value for cell in ws[1]]
            def apply_hyperlink(column_name):
                if column_name not in headers_row:
                    return
                col_idx = headers_row.index(column_name) + 1
                for row_idx in range(2, ws.max_row + 1):
                    cell = ws.cell(row=row_idx, column=col_idx)
                    val = cell.value
                    if val and isinstance(val, str) and val.startswith("http"):
                        cell.hyperlink = val
                        cell.style = "Hyperlink"
            apply_hyperlink(col_image)
            apply_hyperlink(col_url)
            wb.save(excel_path)

    # CSV出力（リンク誤結合/改行対策）
    csv_path = Path(excel_path).with_name("sheet2_llm_targets.csv")
    sheet2_csv = sheet2_df.copy()
    for col in sheet2_csv.columns:
        if sheet2_csv[col].dtype == object:
            sheet2_csv[col] = sheet2_csv[col].astype(str).str.replace("\r\n", " ", regex=False)
            sheet2_csv[col] = sheet2_csv[col].str.replace("\n", " ", regex=False)
            sheet2_csv[col] = sheet2_csv[col].str.replace("\r", " ", regex=False)
    sheet2_csv.to_csv(csv_path, index=False, quoting=csv.QUOTE_ALL, encoding="utf-8", lineterminator="\n")

    if OUTPUT_PAPERS_SHEET2:
        build_papers_sheet2(work, excel_path, target_dates)

    print(f"  Sheet2/CSV 出力: {len(sheet2_df)}件, {csv_path}")

# ========== ソース1: 検証済みRSSフィード ==========
def build_papers_sheet2(df, excel_path, target_dates):
    """Build papers_sheet2.csv for LLM target rows."""
    if df is None or df.empty:
        print("  Papers/CSV: no data")
        return
    col_country = "\u56fd"
    col_title = "\u30bf\u30a4\u30c8\u30eb"
    col_title_jp = "\u30bf\u30a4\u30c8\u30eb\uff08\u65e5\u672c\u8a9e\uff09"
    col_content = "\u5185\u5bb9"
    col_content_jp = "\u5185\u5bb9\uff08\u65e5\u672c\u8a9e\uff09"
    col_site = "\u51fa\u5c55\u30b5\u30a4\u30c8"
    col_image = "\u753b\u50cfURL"
    col_url = "URL"
    col_llm = "LLM\u5224\u5b9a"
    col_date = "\u65e5\u4ed8"

    work = df.copy()
    for col in [col_country, col_title, col_title_jp, col_content, col_content_jp, col_site, col_image, col_url, col_llm, col_date]:
        if col not in work.columns:
            work[col] = ""

    # Fill JP fields if empty
    work[col_title_jp] = work[col_title_jp].where(
        work[col_title_jp].astype(str).str.strip().ne(""), work[col_title]
    )
    work[col_content_jp] = work[col_content_jp].where(
        work[col_content_jp].astype(str).str.strip().ne(""), work[col_content]
    )

    filtered = work[work[col_country] == "\u8ad6\u6587"]
    if target_dates:
        date_set = set(target_dates)
        dt = pd.to_datetime(filtered[col_date], errors="coerce")
        filtered = filtered[dt.dt.strftime("%Y-%m-%d").isin(date_set)]

    filtered = filtered[filtered[col_llm].astype(str).str.strip() == "\u5bfe\u8c61"]

    sheet_cols = [col_country, col_date, col_title_jp, col_content_jp, col_site, col_image, col_url]
    papers_df = filtered[sheet_cols].copy()
    if col_date in papers_df.columns:
        papers_df[col_date] = pd.to_datetime(papers_df[col_date], errors="coerce").dt.strftime("%Y-%m-%d")

    for col in papers_df.columns:
        if papers_df[col].dtype == object:
            papers_df[col] = papers_df[col].astype(str).str.replace("\r\n", " ", regex=False)
            papers_df[col] = papers_df[col].str.replace("\n", " ", regex=False)
            papers_df[col] = papers_df[col].str.replace("\r", " ", regex=False)

    papers_path = Path(excel_path).with_name("papers_sheet2.csv")
    papers_df.to_csv(papers_path, index=False, quoting=csv.QUOTE_ALL, encoding="utf-8-sig", lineterminator="\n")
    print(f"  Papers/CSV output: {len(papers_df)} rows, {papers_path}")

def build_rss_feed_list(path="rss_feed_list.csv"):
    """Write current RSS_FEEDS to csv."""
    rows = []
    for feed in RSS_FEEDS:
        rows.append({
            "\u56fd": feed.get("country", ""),
            "\u51fa\u5c55\u30b5\u30a4\u30c8": feed.get("name", ""),
            "RSS URL": feed.get("url", ""),
        })
    if not rows:
        return
    df_list = pd.DataFrame(rows)
    df_list.to_csv(path, index=False, quoting=csv.QUOTE_ALL, encoding="utf-8-sig", lineterminator="\n")
    print(f"  RSS\u30d5\u30a3\u30fc\u30c9\u4e00\u89a7\u3092\u51fa\u529b: {path}")

def fetch_from_rss(target_dates):
    print("\n=== RSSフィード ===")
    results = []
    
    for feed_info in RSS_FEEDS:
        name = feed_info["name"]
        url = feed_info["url"]
        country = feed_info["country"]
        
        try:
            timeout = 10
            retries = 1
            if country == "論文" or "pubmed.ncbi.nlm.nih.gov" in url:
                timeout = 30
                retries = 2
            response = None
            for _ in range(retries):
                response = requests.get(url, headers=HEADERS, timeout=timeout)
                if response and response.status_code == 200:
                    break
            
            if response and response.status_code == 200:
                feed = feedparser.parse(response.content)
                count = 0
                
                for entry in feed.entries:
                    title = entry.get("title", "")
                    link = entry.get("link", "")
                    published = entry.get("published", "") or entry.get("updated", "")
                    pub_date = parse_date(published)
                    if not pub_date:
                        continue
                    if not is_target_date(pub_date, target_dates):
                        continue
                    
                    desc = ""
                    if 'summary' in entry:
                        desc = BeautifulSoup(str(entry.get("summary", "")), "html.parser").get_text()
                    # Google News??????????URL???????????
                    resolved_link = resolve_final_url(link)
                    # ??URL?RSS?????
                    image_url = extract_image_from_rss(entry)
                    if is_missing_url(image_url):
                        image_url = fetch_image_from_page(resolved_link) if resolved_link else ""
                    if is_missing_url(image_url):
                        # Playwright???URL?????????
                        resolved_pw, image_pw = resolve_with_playwright(resolved_link or link)
                        if resolved_pw:
                            resolved_link = resolved_pw
                        if image_pw:
                            image_url = image_pw
                    # Yahoo yimg placeholder: fallback to og:image
                    if image_url and is_yimg_placeholder(image_url):
                        alt = fetch_image_from_page(resolved_link) if resolved_link else ""
                        if alt:
                            image_url = alt
                    results.append({
                        "国": country,
                        "検索ワード": "RSS",
                        "タイトル": title,
                        "日付": pub_date,
                        "内容": desc[:300] if desc else "",
                        "出展サイト": name,
                        "画像URL": image_url,
                        "URL": link,
                        "ソース": name
                    })
                    count += 1
                
                if count > 0:
                    print(f"  ✓ [{name}] {count}件")
                    
        except KeyboardInterrupt:
            print("  ✗ RSS取得を中断しました。")
            break
        except Exception as e:
            print(f"  ✗ [{name}] エラー")
        
        time.sleep(0.3)
    
    print(f"  RSS合計: {len(results)}件")
    return results

# ========== ソース2: Bing News 検索（Google News代替）==========
def fetch_from_bing_search(target_dates):
    """Bing News RSS検索（キーワード検索対応）"""
    print("\n=== Bing News 検索 ===")
    results = []
    
    for country, settings in COUNTRY_SETTINGS.items():
        market = settings["bing_market"]
        keywords = settings["keywords"]
        
        print(f"  [{country}] 検索中...")
        country_count = 0
        force_skip = False
        
        for keyword in keywords:
            if force_skip:
                break
            encoded_keyword = quote(keyword)
            # Bing側も日付を1日に絞るため when:1d を付与
            rss_url = f"https://www.bing.com/news/search?q={encoded_keyword}%20when%3A1d&format=rss&mkt={market}"
            
            try:
                response = requests.get(rss_url, headers=HEADERS, timeout=10)
                
                if response.status_code == 200:
                    content_type = response.headers.get("Content-Type", "").lower()
                    if ("xml" not in content_type) and (b"<rss" not in response.content[:200]):
                        print(f"    -> {country} / {keyword[:10]}... RSSで返っておらず（HTML/制限の可能性）")
                        if country == "中国":
                            force_skip = True  # 中国マーケットがHTMLなら後続キーワードをスキップ
                        continue
                    feed = feedparser.parse(response.content)
                    
                    for entry in feed.entries[:15]:
                        title = entry.get("title", "")
                        link = entry.get("link", "")
                        published = entry.get("published", "")
                        
                    pub_date = parse_date(published)
                    if not pub_date:
                        continue
                    if not is_target_date(pub_date, target_dates):
                        continue
                    
                    desc = BeautifulSoup(entry.get("summary", ""), "html.parser").get_text()
                    image_url = extract_image_from_rss(entry)
                    # Bingリンクが msn/bing の場合も最終URLを解決して画像再取得
                    resolved_link = resolve_final_url(link)
                    if resolved_link and is_missing_url(image_url):
                        image_url = fetch_image_from_page(resolved_link)
                    
                    results.append({
                        "国": country,
                        "検索ワード": keyword,
                        "タイトル": title,
                        "日付": pub_date,
                        "内容": desc[:300],
                        "出展サイト": get_domain(link),
                        "画像URL": image_url,
                        "URL": resolved_link or link,
                        "ソース": "Bing検索"
                    })
                    country_count += 1
                        
            except:
                pass
            
            time.sleep(0.5)
        
        if country_count > 0:
            print(f"    -> {country_count}件")
        else:
            print("    -> 0件（RSS無し/日付不一致/重複の可能性）")
    
    print(f"  Bing検索合計: {len(results)}件")
    return results

# ========== ソース3: DuckDuckGo News 検索 ==========
def fetch_from_duckduckgo(target_dates):
    """DuckDuckGo News検索"""
    print("\n=== DuckDuckGo News 検索 ===")
    results = []
    
    if not DDGS_AVAILABLE:
        print("  ✗ duckduckgo-search がインストールされていません")
        return results
    
    # 地域設定
    region_map = {
        "日本": "jp-jp",
        "米国": "us-en",
        "インド": "in-en",
        "中国": "cn-zh",
        "欧州": "uk-en"
    }
    
    try:
        ddgs = DDGS()
        
        for country, settings in COUNTRY_SETTINGS.items():
            region = region_map.get(country, "wt-wt")
            # 各国から代表キーワードを拡大して検索件数を増やす
            keywords = settings["keywords"][:10]
            
            print(f"  [{country}] 検索中...")
            country_count = 0
            
            for keyword in keywords:
                try:
                    news_results = ddgs.news(
                        keywords=keyword,
                        region=region,
                        safesearch="off",
                        # 直近1週間を取得し、後段で「昨日」に絞り込む
                        timelimit="w",
                        max_results=25
                    )
                    
                    for item in news_results:
                        title = item.get("title", "")
                        url = item.get("url", "")
                        body = item.get("body", "")
                        date_str = item.get("date", "")
                        source = item.get("source", "")
                        image_url = item.get("image") or item.get("img") or ""
                        
                        pub_date = parse_date(date_str)
                        if not pub_date:
                            # timestamp があれば補完
                            ts = item.get("timestamp")
                            if ts:
                                try:
                                    pub_date = datetime.fromtimestamp(int(ts)).strftime('%Y-%m-%d')
                                except:
                                    pass
                        # 明確に日付が判定できない場合はスキップして精度を担保
                        if not pub_date:
                            continue
                        
                        if not is_target_date(pub_date, target_dates):
                            continue
                        
                        results.append({
                            "国": country,
                            "検索ワード": keyword,
                            "タイトル": title,
                            "日付": pub_date,
                            "内容": body[:300] if body else "",
                            "出展サイト": source,
                            "画像URL": image_url,
                            "URL": url,
                            "ソース": "DuckDuckGo"
                        })
                        country_count += 1
                        
                except Exception as e:
                    pass
                
                time.sleep(1)  # レート制限対策
            
            if country_count > 0:
                print(f"    -> {country_count}件")
            else:
                print("    -> 0件（DuckDuckGoでヒット無し/日付不一致の可能性）")
                
    except Exception as e:
        print(f"  ✗ エラー: {e}")
    
    print(f"  DuckDuckGo合計: {len(results)}件")
    return results

# ========== ソース4: Google News RSS ==========
def fetch_from_google_news(target_dates):
    """Google News RSS（少量・限定キーワードで実行）"""
    print("\n=== Google News RSS ===")
    results = []
    
    google_news_regions = {
        "日本": {"hl": "ja", "gl": "JP", "ceid": "JP:ja"},
        "中国": {"hl": "zh-CN", "gl": "CN", "ceid": "CN:zh-Hans"}
    }
    
    for country, region_params in google_news_regions.items():
        keywords = COUNTRY_SETTINGS.get(country, {}).get("keywords", [])[:3]
        print(f"  [{country}] 検索中...")
        country_count = 0
        
        for keyword in keywords:
            encoded_keyword = quote(keyword)
            rss_url = (
                f"https://news.google.com/rss/search?q={encoded_keyword}%20when:1d"
                f"&hl={region_params['hl']}&gl={region_params['gl']}&ceid={region_params['ceid']}"
            )
            
            try:
                response = requests.get(rss_url, headers=HEADERS, timeout=10)
                if response.status_code != 200:
                    print(f"    -> {keyword[:10]}... 取得失敗 HTTP {response.status_code}")
                    continue
                
                feed = feedparser.parse(response.content)
                for entry in feed.entries[:15]:
                    title = entry.get("title", "")
                    link = entry.get("link", "")
                    published = entry.get("published", "") or entry.get("updated", "")
                    
                    pub_date = parse_date(published)
                    if not pub_date:
                        continue
                    if not is_target_date(pub_date, target_dates):
                        continue
                    
                    desc = BeautifulSoup(str(entry.get("summary", "")), "html.parser").get_text()
                    # Google News??????????URL???????????
                    resolved_link = resolve_final_url(link)
                    # ??URL?RSS?????
                    image_url = extract_image_from_rss(entry)
                    if is_missing_url(image_url):
                        image_url = fetch_image_from_page(resolved_link) if resolved_link else ""
                    if is_missing_url(image_url):
                        # Playwright???URL?????????
                        resolved_pw, image_pw = resolve_with_playwright(resolved_link or link)
                        if resolved_pw:
                            resolved_link = resolved_pw
                        if image_pw:
                            image_url = image_pw
                    results.append({
                        "国": country,
                        "検索ワード": keyword,
                        "タイトル": title,
                        "日付": pub_date,
                        "内容": desc[:300] if desc else "",
                        "出展サイト": get_domain(link),
                        "画像URL": image_url,
                        "URL": resolved_link or link,
                        "ソース": "GoogleNews"
                    })
                    country_count += 1
            except Exception as e:
                print(f"    -> {keyword[:10]}... 取得エラー {e}")
            
            time.sleep(0.5)
        
        if country_count > 0:
            print(f"    -> {country_count}件")
        else:
            print("    -> 0件（ヒット無し/日付不一致/重複の可能性）")
    
    print(f"  Google News合計: {len(results)}件")
    return results

# ========== ソース5: NewsAPI ==========
def fetch_from_newsapi(target_dates):
    print("\n=== NewsAPI.org ===")
    results = []
    
    if target_dates:
        try:
            dates_dt = [datetime.strptime(d, "%Y-%m-%d") for d in target_dates]
            from_date = min(dates_dt).strftime("%Y-%m-%d")
            to_date = max(dates_dt).strftime("%Y-%m-%d")
        except Exception:
            from_date = (datetime.now() - timedelta(days=1)).strftime('%Y-%m-%d')
            to_date = datetime.now().strftime('%Y-%m-%d')
    else:
        from_date = (datetime.now() - timedelta(days=1)).strftime('%Y-%m-%d')
        to_date = datetime.now().strftime('%Y-%m-%d')
    
    # 各国から代表キーワードを選択（API制限のため少なめ）
    keywords_to_search = []
    for country, settings in COUNTRY_SETTINGS.items():
        lang = settings["lang"]
        for kw in settings["keywords"][:2]:
            keywords_to_search.append((kw, lang, country))
    
    for keyword, lang, country in keywords_to_search:
        params = {
            "q": keyword,
            "from": from_date,
            "to": to_date,
            "language": lang,
            "sortBy": "publishedAt",
            "pageSize": 30,
            "apiKey": NEWSAPI_KEY
        }
        
        try:
            response = requests.get(NEWSAPI_ENDPOINT, params=params, timeout=15)
            
            if response.status_code == 200:
                data = response.json()
                articles = data.get("articles", [])
                count = 0
                
                for article in articles:
                    pub_date = article.get("publishedAt", "")[:10]
                    if not is_target_date(pub_date, target_dates):
                        continue
                    image_url = article.get("urlToImage", "")
                        
                    results.append({
                        "国": country,
                        "検索ワード": keyword,
                        "タイトル": article.get("title", ""),
                        "日付": pub_date,
                        "内容": article.get("description", "") or "",
                        "出展サイト": article.get("source", {}).get("name", ""),
                        "画像URL": image_url,
                        "URL": article.get("url", ""),
                        "ソース": "NewsAPI"
                    })
                    count += 1
                
                if count > 0:
                    print(f"  ✓ [{country}] {keyword[:15]}...: {count}件")
                    
            elif response.status_code == 429:
                print(f"  ✗ レート制限到達")
                break
                
        except:
            pass
        
        time.sleep(0.5)
    
    print(f"  NewsAPI合計: {len(results)}件")
    return results

def enrich_results(items, label="新規", existing_df=None, save_path=None):
    """関連度スコアと日本語要約を付与する"""
    total = len(items)
    if total:
        print(f"  [{label}] 関連度・翻訳処理 {total}件 開始")
    missing_img_urls = []
    for idx, item in enumerate(items, 1):
        title = item.get("タイトル", "")
        content = str(item.get("内容", "")).strip()
        item["内容"] = content
        score, rel_label, hits = compute_relevance(title, content)
        item["関連度スコア"] = score
        item["関連度"] = rel_label
        item["関連キーワード"] = ", ".join(hits)
        item.setdefault("HTML取得", "可能")
        item.setdefault("LLM判定", "")
        item.setdefault("画像判定", "")
        item.setdefault("LLM後処理", "")
        
        country = item.get("国", "")
        source = item.get("ソース", "")
        url_value = item.get("URL")
        # LLM判定（関連性のみ）を先に実施
        llm_relevance = ""
        if USE_LLM:
            llm_relevance, _ = call_llm_classify(title, content, "", mode="relevance")
            if llm_relevance:
                item["LLM判定"] = llm_relevance
        # GoogleNewsのURLは強制的に解決を試みる
        if source == "GoogleNews" and isinstance(url_value, str) and "news.google.com" in url_value:
            resolved_pw, image_pw = resolve_with_playwright(url_value)
            if resolved_pw:
                item["URL"] = resolved_pw
            if image_pw and is_missing_url(item.get("????RL")):
                item["画像URL"] = image_pw
            url_value = item.get("URL")
        # なお残る news.google.com や空欄はタイトル検索で補完
        if source == "GoogleNews" and (not item.get("URL") or (isinstance(item.get("URL"), str) and "news.google.com" in item.get("URL"))):
            guessed = search_article_url(title, country)
            if guessed:
                item["URL"] = guessed
                if is_missing_url(item.get("画像URL")):
                    item["画像URL"] = fetch_image_from_page(guessed)
        if FETCH_MISSING_IMAGES and is_missing_url(item.get("画像URL")) and item.get("URL"):
            resolved = resolve_final_url(item.get("URL"))
            item["URL"] = resolved
            missing_img_urls.append(resolved)
            if is_missing_url(item.get("画像URL")):
                resolved_pw, image_pw = resolve_with_playwright(resolved)
                if resolved_pw:
                    item["URL"] = resolved_pw
                if image_pw:
                    item["画像URL"] = image_pw
        # URL健全性チェック（記事）
        if not check_url_ok(item.get("URL"), is_image=False):
            item["HTML取得"] = "×"
        # 画像の健全性チェック
        if not is_missing_url(item.get("画像URL")):
            if not check_url_ok(item.get("画像URL"), is_image=True):
                item["画像URL"] = ""
        # まだ news.google.com のままなら無効として空欄
        if source == "GoogleNews" and isinstance(item.get("URL"), str) and "news.google.com" in item.get("URL"):
            item["HTML取得"] = "×"
            item["URL"] = ""
            item["画像URL"] = ""
        # LLM判定が非対象なら後処理をスキップ
        if USE_LLM and item.get("LLM判定") == "非対象":
            item["LLM後処理"] = "スキップ"
        else:
            # 要約（日本語）を生成
            summary_title, summary_body = summarize_article(title, content, item.get("URL", ""), country)
            if summary_title:
                item["タイトル（日本語）"] = summary_title
            if summary_body:
                item["内容（日本語）"] = summary_body
            # 画像判定（対象のみ）
            if USE_LLM:
                _, llm_photo = call_llm_classify(title, content, item.get("画像URL", ""), mode="photo")
                if llm_photo:
                    item["画像判定"] = llm_photo
            item["LLM後処理"] = "実施"
        if total and (idx == 1 or idx % PROGRESS_EVERY == 0 or idx == total):
            print(f"    進捗: {idx}/{total}")
        if LLM_SAVE_INTERVAL > 0 and idx % LLM_SAVE_INTERVAL == 0 and existing_df is not None and save_path:
            try:
                partial_df = pd.DataFrame(items[:idx])
                for col in OUTPUT_COLUMNS:
                    if col not in partial_df.columns:
                        partial_df[col] = ""
                partial_df = partial_df[OUTPUT_COLUMNS]
                if not existing_df.empty:
                    for col in partial_df.columns:
                        if col not in existing_df.columns:
                            existing_df[col] = ""
                    merged = pd.concat([partial_df, existing_df], ignore_index=True)
                else:
                    merged = partial_df
                save_with_hyperlinks(merged, save_path)
                print(f"    途中保存: {idx}件")
            except Exception:
                pass
    if total:
        print(f"  [{label}] 関連度・翻訳処理 完了")
        if FETCH_MISSING_IMAGES and missing_img_urls:
            print(f"    画像補完: {len(missing_img_urls)}件 並列取得開始")
            img_map = bulk_fetch_images(missing_img_urls)
            for item in items:
                if is_missing_url(item.get("画像URL")) and item.get("URL"):
                    item["画像URL"] = img_map.get(item.get("URL"), item.get("画像URL"))
            empty_after = sum(1 for it in items if is_missing_url(it.get("画像URL")))
            print(f"    画像URL未取得: {empty_after}件")
    return items

def enrich_existing_df(df):
    """既存データにも関連度と日本語要約を付与する"""
    if df.empty:
        return df
    for col, default in [
        ("関連度", ""),
        ("関連度スコア", 0.0),
        ("関連キーワード", ""),
        ("タイトル（日本語）", ""),
        ("内容（日本語）", ""),
        ("画像URL", ""),
        ("HTML取得", "可能"),
        ("LLM判定", ""),
        ("画像判定", ""),
        ("LLM後処理", "")
    ]:
        if col not in df.columns:
            df[col] = default
    total = len(df)
    df_out = df.copy()
    print(f"  [既存] 関連度・翻訳再計算 {total}件 開始")
    if USE_LLM:
        print("  [既存] LLM判定: 有効")
    if LLM_ONLY:
        print("  [既存] LLM-only: URL/翻訳/画像チェックを省略")
    missing_img_urls = []
    llm_calls = 0
    for count, (row_idx, row) in enumerate(df_out.iterrows(), 1):
        title = row.get("タイトル", "")
        content = str(row.get("内容", "")).strip()
        row["内容"] = content
        score, label, hits = compute_relevance(title, content)
        row["関連度スコア"] = score
        row["関連度"] = label
        row["関連キーワード"] = ", ".join(hits)
        if "HTML取得" not in row or pd.isna(row.get("HTML取得")):
            row["HTML取得"] = "可能"
        if "LLM判定" not in row or pd.isna(row.get("LLM判定")):
            row["LLM判定"] = ""
        if "画像判定" not in row or pd.isna(row.get("画像判定")):
            row["画像判定"] = ""
        if RESUME_LLM and str(row.get("LLM後処理")).strip() in ("実施", "スキップ"):
            df_out.loc[row_idx] = row
            if total and (count == 1 or count % PROGRESS_EVERY == 0 or count == total):
                print(f"    進捗: {count}/{total}")
            continue
        if LLM_ONLY:
            if USE_LLM:
                llm_need = False
                if pd.isna(row.get("LLM判定")) or str(row.get("LLM判定")).strip() == "":
                    llm_need = True
                if pd.isna(row.get("画像判定")) or str(row.get("画像判定")).strip() == "":
                    llm_need = True
                if llm_need:
                    llm_rel, _ = call_llm_classify(title, content, "", mode="relevance")
                    if llm_rel:
                        row["LLM判定"] = llm_rel
                    if row.get("LLM判定") == "対象":
                        _, llm_photo = call_llm_classify(title, content, row.get("画像URL", ""), mode="photo")
                        if llm_photo:
                            row["画像判定"] = llm_photo
                        row["LLM後処理"] = "実施"
                    elif row.get("LLM判定") == "非対象":
                        row["LLM後処理"] = "スキップ"
                    llm_calls += 1
            df_out.loc[row_idx] = row
            if total and (count == 1 or count % PROGRESS_EVERY == 0 or count == total):
                print(f"    進捗: {count}/{total}")
            if USE_LLM and ENRICH_ONLY and LLM_SAVE_INTERVAL > 0 and count % LLM_SAVE_INTERVAL == 0:
                try:
                    temp_df = df_out.copy()
                    save_with_hyperlinks(temp_df, EXCEL_FILE)
                    print(f"    途中保存: {count}件")
                except Exception:
                    pass
            continue
        source = row.get("ソース", "")
        url_value = row.get("URL")
        # GoogleNewsのURLは強制的に解決を試みる
        if source == "GoogleNews" and isinstance(url_value, str) and "news.google.com" in url_value:
            resolved_pw, image_pw = resolve_with_playwright(url_value)
            if resolved_pw:
                row["URL"] = resolved_pw
            if image_pw and is_missing_url(row.get("????RL")):
                row["画像URL"] = image_pw
        if FETCH_MISSING_IMAGES and is_missing_url(row.get("画像URL")) and row.get("URL"):
            resolved = resolve_final_url(row.get("URL"))
            row["URL"] = resolved
            missing_img_urls.append(resolved)
            if is_missing_url(row.get("画像URL")):
                resolved_pw, image_pw = resolve_with_playwright(resolved)
                if resolved_pw:
                    row["URL"] = resolved_pw
                if image_pw:
                    row["画像URL"] = image_pw
        # なお残る news.google.com や空欄はタイトル検索で補完
        if source == "GoogleNews" and (not row.get("URL") or (isinstance(row.get("URL"), str) and "news.google.com" in row.get("URL"))):
            guessed = search_article_url(title, row.get("国", ""))
            if guessed:
                row["URL"] = guessed
                if is_missing_url(row.get("画像URL")):
                    row["画像URL"] = fetch_image_from_page(guessed)
        # URL健全性チェック（記事）
        if not check_url_ok(row.get("URL"), is_image=False):
            row["HTML取得"] = "×"
        # 画像の健全性チェック
        if not is_missing_url(row.get("画像URL")):
            if not check_url_ok(row.get("画像URL"), is_image=True):
                row["画像URL"] = ""
        # まだ news.google.com のままなら無効として空欄
        if source == "GoogleNews" and isinstance(row.get("URL"), str) and "news.google.com" in row.get("URL"):
            row["HTML取得"] = "×"
            row["URL"] = ""
            row["画像URL"] = ""
        force_process = PROCESS_LLM_SKIPPED and str(row.get("LLM後処理")).strip() == "スキップ"
        if USE_LLM and not force_process:
            if pd.isna(row.get("LLM判定")) or str(row.get("LLM判定")).strip() == "":
                llm_rel, _ = call_llm_classify(title, content, "", mode="relevance")
                if llm_rel:
                    row["LLM判定"] = llm_rel
                llm_calls += 1
            if row.get("LLM判定") == "非対象":
                row["LLM後処理"] = "スキップ"
                df_out.loc[row_idx] = row
                if total and (count == 1 or count % PROGRESS_EVERY == 0 or count == total):
                    print(f"    進捗: {count}/{total}")
                if USE_LLM and ENRICH_ONLY and LLM_SAVE_INTERVAL > 0 and count % LLM_SAVE_INTERVAL == 0:
                    try:
                        temp_df = df_out.copy()
                        save_with_hyperlinks(temp_df, EXCEL_FILE)
                        print(f"    途中保存: {count}件")
                    except Exception:
                        pass
                continue
        # 要約（日本語）を生成
        summary_title, summary_body = summarize_article(title, content, row.get("URL", ""), row.get("国", ""))
        if summary_title:
            row["タイトル（日本語）"] = summary_title
        if summary_body:
            row["内容（日本語）"] = summary_body
        # 画像判定（対象/強制処理のみ）
        if USE_LLM:
            _, llm_photo = call_llm_classify(title, content, row.get("画像URL", ""), mode="photo")
            if llm_photo:
                row["画像判定"] = llm_photo
            llm_calls += 1
        row["LLM後処理"] = "実施"
        df_out.loc[row_idx] = row
        if total and (count == 1 or count % PROGRESS_EVERY == 0 or count == total):
            print(f"    進捗: {count}/{total}")
        if USE_LLM and ENRICH_ONLY and LLM_SAVE_INTERVAL > 0 and count % LLM_SAVE_INTERVAL == 0:
            try:
                temp_df = df_out.copy()
                save_with_hyperlinks(temp_df, EXCEL_FILE)
                print(f"    途中保存: {count}件")
            except Exception:
                pass
    print("  [既存] 関連度・翻訳再計算 完了")
    if USE_LLM:
        print(f"  [既存] LLM判定 実行件数: {llm_calls}")
    if FETCH_MISSING_IMAGES:
        if missing_img_urls:
            print(f"    画像補完: {len(missing_img_urls)}件 並列取得開始")
            img_map = bulk_fetch_images(missing_img_urls)
            for r_idx, r in df_out.iterrows():
                if is_missing_url(r.get("画像URL")) and r.get("URL"):
                    df_out.at[r_idx, "画像URL"] = img_map.get(r.get("URL"), r.get("画像URL"))
        empty_after = sum(1 for _, r in df_out.iterrows() if is_missing_url(r.get("画像URL")))
        print(f"    画像URL未取得: {empty_after}件（補完試行: {len(missing_img_urls)}件）")
    return df_out

def main():
    os.chdir(os.path.dirname(os.path.abspath(__file__)))
    if "--dept" in sys.argv:
        try:
            idx = sys.argv.index("--dept")
            if len(sys.argv) > idx + 1:
                global DEPARTMENT
                DEPARTMENT = sys.argv[idx + 1]
        except Exception:
            pass
    if not DEPARTMENT and Path(DEPT_SETTINGS_PATH).exists():
        DEPARTMENT = "interior"
    apply_department_settings(DEPARTMENT, DEPT_SETTINGS_PATH)



    today = datetime.now()
    yesterday = today - timedelta(days=1)
    target_dates = []
    if "--dates" in sys.argv:
        try:
            idx = sys.argv.index("--dates")
            dates_arg = sys.argv[idx + 1]
            target_dates = [d.strip() for d in dates_arg.split(",") if d.strip()]
        except Exception:
            target_dates = []
    elif os.getenv("TARGET_DATES"):
        target_dates = [d.strip() for d in os.getenv("TARGET_DATES", "").split(",") if d.strip()]
    elif TARGET_DATE_ONLY_YESTERDAY:
        target_dates = [yesterday.strftime('%Y-%m-%d')]

    if ONLY_PAPERS_RSS and not target_dates:
        target_dates = []
    
    print("=" * 60)
    print(f"  {SUBJECT_NAME}ニュース収集スクリプト（検証済みRSS版）")
    print("=" * 60)
    if target_dates:
        print(f"対象日付: {', '.join(target_dates)}")
    else:
        print("対象日付: 制限なし（全期間）")
    print(f"RSSフィード: {len(RSS_FEEDS)}サイト（検証済み）")
    build_rss_feed_list()
    print(f"Bing検索: 各国{len(COUNTRY_SETTINGS)}地域 × 16-18キーワード")

    # 既存データ読み込み
    existing_urls = set()
    collected_titles = []
    df_existing = pd.DataFrame()

    if os.path.exists(EXCEL_FILE):
        df_existing = load_existing_data(EXCEL_FILE)
    elif os.path.exists(LEGACY_EXCEL_FILE):
        df_existing = load_existing_data(LEGACY_EXCEL_FILE)

    if not df_existing.empty:
        if "URL" in df_existing.columns:
            existing_urls = set(
                u for u in df_existing["URL"].tolist() if str(u).strip()
            )
        if "タイトル" in df_existing.columns:
            collected_titles = [
                str(t) for t in df_existing["タイトル"].tolist() if str(t).strip()
            ]
        if ENRICH_ONLY or ENRICH_EXISTING:
            df_existing = enrich_existing_df(df_existing)
            print(f"既存データ: {len(df_existing)} 件")

    # 既存のみ再計算モード
    if ENRICH_ONLY:
        if df_existing.empty:
            print("既存データがありません。処理を終了します。")
            return
        for col in OUTPUT_COLUMNS:
            if col not in df_existing.columns:
                df_existing[col] = ""
        extra_cols = [c for c in df_existing.columns if c not in OUTPUT_COLUMNS]
        df_final = df_existing[OUTPUT_COLUMNS + extra_cols]
        try:
            save_with_hyperlinks(df_final, EXCEL_FILE)
            print(f"\n✅ 完了！ 既存 {len(df_existing)} 件を再計算し保存しました。")
            print(f"   保存先: {os.path.abspath(EXCEL_FILE)}")
            build_sheet2_and_csv(df_final, EXCEL_FILE, target_dates)
        except Exception as e:
            print(f"\n❌ 保存エラー: {e}")
        return

    all_results = []

    if ONLY_PAPERS_RSS:
        RSS_FEEDS[:] = [f for f in RSS_FEEDS if f.get("country") == "論文"]
    
    # 各ソースから取得
    all_results.extend(fetch_from_rss(target_dates))
    if not ONLY_PAPERS_RSS:
        all_results.extend(fetch_from_bing_search(target_dates))
        all_results.extend(fetch_from_duckduckgo(target_dates))
        if ENABLE_GOOGLE_NEWS:
            all_results.extend(fetch_from_google_news(target_dates))
        all_results.extend(fetch_from_newsapi(target_dates))
    
    print("\n" + "=" * 60)
    print(f"全ソース合計: {len(all_results)}件")
    if all_results:
        src_counter = Counter([item.get("ソース", "") for item in all_results if item.get("ソース")])
        cty_counter = Counter([item.get("国", "") for item in all_results if item.get("国")])
        print(f"  取得内訳（ソース別）: {dict(src_counter)}")
        print(f"  取得内訳（国別）: {dict(cty_counter)}")
    
    # 重複除去
    print("\n=== 重複除去処理 ===")
    unique_results = []
    seen_urls = set(existing_urls)
    seen_titles = list(collected_titles)
    dup_by_source = Counter()
    dup_by_reason = Counter()
    
    for item in all_results:
        url = item.get("URL", "")
        title = item.get("タイトル", "")
        source_name = item.get("ソース", "不明")
        
        if not url or not title:
            continue
        
        if url in seen_urls:
            dup_by_source[source_name] += 1
            dup_by_reason["既存URL"] += 1
            continue
        
        is_dup = False
        for t in seen_titles[-500:]:
            if is_similar(title, t):
                is_dup = True
                break
        if is_dup:
            dup_by_source[source_name] += 1
            dup_by_reason["タイトル類似"] += 1
            continue
        
        seen_urls.add(url)
        seen_titles.append(title)
        unique_results.append(item)
    
    removed = len(all_results) - len(unique_results)
    print(f"  重複除去: {removed}件")
    if dup_by_source:
        print(f"  重複内訳（ソース別）: {dict(dup_by_source)}")
    if dup_by_reason:
        print(f"  重複理由内訳: {dict(dup_by_reason)}")
    print(f"  最終結果: {len(unique_results)}件")
    if unique_results:
        src_counter_final = Counter([item.get("ソース", "") for item in unique_results if item.get("ソース")])
        cty_counter_final = Counter([item.get("国", "") for item in unique_results if item.get("国")])
        print(f"  最終内訳（ソース別）: {dict(src_counter_final)}")
        print(f"  最終内訳（国別）: {dict(cty_counter_final)}")
    
    if unique_results:
        unique_results = enrich_results(unique_results, label="新規", existing_df=df_existing, save_path=EXCEL_FILE)
        df_new = pd.DataFrame(unique_results)
        
        for col in OUTPUT_COLUMNS:
            if col not in df_new.columns:
                df_new[col] = ""
        df_new = df_new[OUTPUT_COLUMNS]
        
        df_new["ステータス"] = "OK"
        df_new["HTML取得"] = "可能"
        
        if not df_existing.empty:
            for col in df_new.columns:
                if col not in df_existing.columns:
                    df_existing[col] = ""
            df_final = pd.concat([df_new, df_existing], ignore_index=True)
        else:
            df_final = df_new
            
        try:
            save_with_hyperlinks(df_final, EXCEL_FILE)
            print(f"\n✅ 完了！ {len(unique_results)} 件の新規記事を追加しました。")
            print(f"   URLはハイパーリンク化済み")
            print(f"   保存先: {os.path.abspath(EXCEL_FILE)}")
            build_sheet2_and_csv(df_final, EXCEL_FILE, target_dates)
        except Exception as e:
            print(f"\n❌ 保存エラー: {e}")
    else:
        print("\n新規記事は見つかりませんでした。")
        if not df_existing.empty:
            try:
                build_sheet2_and_csv(df_existing, EXCEL_FILE, target_dates)
            except Exception:
                pass

if __name__ == "__main__":
    main()
